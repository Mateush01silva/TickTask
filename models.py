
import pandas as pd
import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

class TaskTickModel:
    def __init__(self, db_path):
        self.db_path = db_path

    def _criar_abas_iniciais(self):
        """
        Cria as abas "Atividades" e "Horas Trabalhadas" se não existirem no arquivo Excel.
        """
        with pd.ExcelWriter(self.db_path, engine='openpyxl', mode='a' if os.path.exists(self.db_path) else 'w') as writer:
            if 'Atividades' not in writer.book.sheetnames:
                df_atividades = pd.DataFrame(columns=['Projeto', 'Descrição', 'Cliente'])
                df_atividades.to_excel(writer, sheet_name='Atividades', index=False)
            if 'Horas Trabalhadas' not in writer.book.sheetnames:
                df_horas = pd.DataFrame(columns=['Atividade', 'Data', 'Hora Início', 'Hora Fim'])
                df_horas.to_excel(writer, sheet_name='Horas Trabalhadas', index=False)

    def carregar_atividades(self):
        if not os.path.exists(self.db_path):
            self._criar_abas_iniciais()
            return []
        df = pd.read_excel(self.db_path, sheet_name="Atividades")
        return df['Projeto'].tolist()
        
    def cadastrar_atividade(self, projeto, descricao, cliente):
        try:
            if not os.path.exists(self.db_path):
                self._criar_abas_iniciais()

            df = pd.read_excel(self.db_path, sheet_name='Atividades')

            proxima_linha = df.shape[0]

            df.loc[proxima_linha] = [projeto, descricao, cliente]

            with pd.ExcelWriter(self.db_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name='Atividades', index=False)
            
        except PermissionError:
            raise PermissionError("O arquivo já está aberto. Por favor, feche o arquivo e tente novamente.")

    def registrar_horas(self, atividade, data, hora_inicio, hora_fim):
        try:
            if not os.path.exists(self.db_path):
                self._criar_abas_iniciais()

            if 'Horas Trabalhadas' in pd.ExcelFile(self.db_path).sheet_names:
                df = pd.read_excel(self.db_path, sheet_name='Horas Trabalhadas')
            else:
                df = pd.DataFrame(columns=['Atividade', 'Data', 'Hora Início', 'Hora Fim'])

            proxima_linha = df.shape[0]

            df.loc[proxima_linha] = [atividade, data, hora_inicio, hora_fim]

            with pd.ExcelWriter(self.db_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name='Horas Trabalhadas', index=False)

        except PermissionError:
            raise PermissionError("O arquivo já está aberto. Por favor, feche o arquivo e tente novamente.")

""" import pandas as pd
import os
from openpyxl import load_workbook
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

class TaskTickModel:
    def __init__(self, db_path):
        self.db_path = db_path

    def carregar_atividades(self):
        if not os.path.exists(self.db_path):
            return []
        df = pd.read_excel(self.db_path, sheet_name="Atividades")
        return df['Projeto'].tolist()
        
    def cadastrar_atividade(self, projeto, descricao, cliente):
            try:
                df = pd.read_excel(self.db_path, sheet_name='Atividades')

                proxima_linha = df.shape[0] + 1

                df.loc[proxima_linha, 'Projeto'] = projeto
                df.loc[proxima_linha, 'Descrição'] = descricao
                df.loc[proxima_linha, 'Cliente'] = cliente

                df.to_excel(self.db_path, index=False, sheet_name='Atividades')
            
            except PermissionError:
                raise PermissionError("O arquivo já está aberto. Por favor, feche o arquivo e tente novamente.")

    def registrar_horas(self, atividade, data, hora_inicio, hora_fim):
        try:
            # Verificar se o arquivo e a aba existem
            if os.path.exists(self.db_path):
                book = load_workbook(self.db_path)
                if 'Horas Trabalhadas' in book.sheetnames:
                    df = pd.read_excel(self.db_path, sheet_name='Horas Trabalhadas')
                else:
                    # Criar um novo dataframe se a aba não existir
                    df = pd.DataFrame(columns=['Atividade', 'Data', 'Hora Início', 'Hora Fim'])
            else:
                # Criar um novo dataframe se o arquivo não existir
                df = pd.DataFrame(columns=['Atividade', 'Data', 'Hora Início', 'Hora Fim'])

            # Obter o número da próxima linha disponível (última linha + 1)
            proxima_linha = df.shape[0]

            # Adicionar os valores nas colunas especificadas
            df.loc[proxima_linha] = [atividade, data, hora_inicio, hora_fim]

            # Salvar de volta no arquivo Excel
            #with pd.ExcelWriter(self.db_path, engine='openpyxl', mode='a') as writer:
            #    writer.book = book
            df.to_excel(self.db_path, index=False, sheet_name='Horas Trabalhadas')

        except PermissionError:
            raise PermissionError("O arquivo já está aberto. Por favor, feche o arquivo e tente novamente.") """
